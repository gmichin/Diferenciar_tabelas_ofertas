import pdfplumber
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def sanitize_sheet_name(name):
    """Remove caracteres inválidos para nomes de aba do Excel"""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name[:31]  # Limita a 31 caracteres

def extract_date_from_pdf(pdf_path):
    """Extrai a data do texto do PDF"""
    date_pattern = r'\d{2}\s?[/-]\s?\d{2}\s?[/-]\s?\d{2,4}|\d{1,2}\s?de\s?[A-Za-zç]+\s?de\s?\d{4}'
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            
            # Procura acima da LEGENDA
            if 'LEGENDA' in text:
                legenda_index = text.index('LEGENDA')
                search_area = text[:legenda_index]
                date_match = re.search(date_pattern, search_area)
                if date_match:
                    return sanitize_sheet_name(date_match.group())
            
            # Procura no canto direito
            right_side = text[-200:]
            date_match = re.search(date_pattern, right_side)
            if date_match:
                return sanitize_sheet_name(date_match.group())
    
    return "Data_nao_encontrada"

def process_pdf_to_dataframe(pdf_path):
    """Processa um único PDF e retorna um DataFrame e a data"""
    filename = os.path.basename(pdf_path).replace('.pdf', '')
    date = extract_date_from_pdf(pdf_path)
    
    # Extrair todas as tabelas
    all_tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                all_tables.extend(table)
    
    if not all_tables:
        print(f"Nenhuma tabela encontrada em {pdf_path}.")
        return None, None, None
    
    # Encontrar cabeçalho - procura a linha com "CÓD. REF"
    headers = ["CÓD. REF", "MARCA", "PRODUTO", "PESO CAIXA", 
              "VALOR 3%", "VALOR 1%", "STATUS", "ESTOQUE", "CX's"]
    
    # Encontrar a linha de cabeçalho real
    header_row_idx = None
    for i, row in enumerate(all_tables):
        if any("CÓD. REF" in str(cell) for cell in row):
            header_row_idx = i
            break
    
    if header_row_idx is None:
        print(f"Cabeçalho não encontrado em {pdf_path}, usando padrão")
        header_row_idx = 0
        actual_headers = all_tables[header_row_idx]
    else:
        actual_headers = all_tables[header_row_idx]
    
    # Coletar todas as linhas abaixo do cabeçalho
    data_rows = []
    for row in all_tables[header_row_idx+1:]:
        if len(row) == len(actual_headers):  # Só adiciona se tiver número correto de colunas
            data_rows.append(row)
    
    # Criar DataFrame
    df = pd.DataFrame(data_rows, columns=actual_headers)
    
    # Limpeza básica - remover linhas vazias
    df = df[df.iloc[:, 0].str.strip().astype(bool)]
    
    return df, date, filename

def compare_dataframes(df1, df2, date1, date2):
    """Compara dois dataframes e retorna um dataframe com as diferenças"""
    # Criar cópias para não modificar os originais
    df1 = df1.copy()
    df2 = df2.copy()
    
    # Padronizar colunas de código de referência
    df1['COD_REF'] = df1.iloc[:, 0].str.strip()
    df2['COD_REF'] = df2.iloc[:, 0].str.strip()
    
    # Encontrar todos os códigos únicos
    all_codes = set(df1['COD_REF']).union(set(df2['COD_REF']))
    
    differences = []
    red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
    
    for code in all_codes:
        row1 = df1[df1['COD_REF'] == code]
        row2 = df2[df2['COD_REF'] == code]
        
        # Caso 1: código existe apenas em um dataframe
        if row1.empty or row2.empty:
            source_df = df1 if row2.empty else df2
            source_date = date1 if row2.empty else date2
            row = source_df[source_df['COD_REF'] == code].iloc[0].copy()
            row['ORIGEM'] = f"Exclusivo em {source_date}"
            differences.append((row.to_dict(), True))  # True = linha toda destacada
            continue
        
        # Caso 2: código existe em ambos, mas há diferenças
        row1 = row1.iloc[0]
        row2 = row2.iloc[0]
        
        different_columns = []
        merged_row = row1.copy()
        
        for col in df1.columns:
            if col == 'COD_REF':
                continue
                
            val1 = str(row1[col]).strip()
            val2 = str(row2[col]).strip()
            
            if val1 != val2:
                different_columns.append(col)
                merged_row[col] = f"{val1}/{val2}"
        
        if different_columns:
            merged_row['ORIGEM'] = f"Diferença em {', '.join(different_columns)}"
            differences.append((merged_row.to_dict(), different_columns))
    
    if not differences:
        return None
    
    # Criar dataframe de diferenças
    diff_df = pd.DataFrame([diff[0] for diff in differences])
    
    # Remover coluna auxiliar COD_REF
    diff_df = diff_df.drop(columns=['COD_REF'], errors='ignore')
    
    return diff_df, [(diff[1], idx) for idx, diff in enumerate(differences)]

def create_excel_table(ws, df, start_row, start_col=1, table_name=None):
    """Cria uma tabela formatada no Excel com filtros sem duplicar cabeçalho"""
    # Escrever apenas os dados (sem cabeçalho)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Adicionar cabeçalho manualmente com formatação
    header_font = Font(bold=True)
    for c_idx, column_name in enumerate(df.columns, start_col):
        cell = ws.cell(row=start_row-1, column=c_idx, value=column_name)
        cell.font = header_font
    
    # Determinar o intervalo da tabela (incluindo cabeçalho)
    max_row = start_row + len(df) - 1
    max_col = start_col + len(df.columns) - 1
    
    # Criar o nome da tabela se não fornecido
    if table_name is None:
        table_name = f"Table_{ws.title}_{start_row}"
        table_name = re.sub(r'\W+', '_', table_name)
    
    # Criar o objeto Table (incluindo cabeçalho na referência)
    tab = Table(displayName=table_name, 
               ref=f"{get_column_letter(start_col)}{start_row-1}:{get_column_letter(max_col)}{max_row}")
    
    # Adicionar um estilo padrão com filtros
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                         showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    
    # Adicionar a tabela à planilha
    ws.add_table(tab)
    
    # Ajustar largura das colunas
    for col_idx in range(start_col, max_col + 1):
        max_length = max(
            len(str(df.columns[col_idx-start_col])),  # Tamanho do cabeçalho
            df.iloc[:, col_idx-start_col].astype(str).str.len().max()  # Tamanho dos dados
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    return ws

def pdfs_to_excel_with_sheets(pdf_paths, output_excel_path=None):
    """Converte múltiplos PDFs para um único Excel com abas diferentes"""
    try:
        if not pdf_paths:
            print("Nenhum arquivo PDF fornecido.")
            return
        
        if output_excel_path is None:
            output_excel_path = os.path.join(os.path.dirname(pdf_paths[0]), "Tabelas_Consolidadas.xlsx")
        
        # Criar um writer Excel
        writer = pd.ExcelWriter(output_excel_path, engine='openpyxl')
        writer.book = load_workbook(output_excel_path) if os.path.exists(output_excel_path) else writer.book
        
        # Processar todos os PDFs primeiro
        processed_data = []
        for pdf_path in pdf_paths:
            df, date, filename = process_pdf_to_dataframe(pdf_path)
            if df is not None:
                processed_data.append((df, date, filename, pdf_path))
        
        # Escrever cada PDF em uma aba separada
        for df, date, filename, pdf_path in processed_data:
            sheet_name = date if date else sanitize_sheet_name(filename)
            
            # Se a data já existir como aba, adiciona um sufixo
            original_sheet_name = sheet_name
            counter = 1
            while sheet_name in writer.book.sheetnames:
                sheet_name = f"{original_sheet_name}_{counter}"
                counter += 1
            
            # Criar nova aba
            ws = writer.book.create_sheet(sheet_name)
            
            # Adicionar título mesclado
            ws['A1'] = f"{filename} - {date.replace('_', ' ') if date else filename}"
            ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
            
            # Configurar estilo do título
            title_font = Font(bold=True, size=12)
            title_alignment = Alignment(horizontal='center')
            ws['A1'].font = title_font
            ws['A1'].alignment = title_alignment
            
            # Criar tabela formatada com filtros (começa na linha 3)
            create_excel_table(ws, df, start_row=3, table_name=f"Tabela_{sheet_name}")
            
            print(f"Processado: {pdf_path} -> aba '{sheet_name}' ({len(df)} linhas)")
        
        # Adicionar aba de diferenças se houver exatamente 2 PDFs
        if len(processed_data) == 2:
            df1, date1, filename1, _ = processed_data[0]
            df2, date2, filename2, _ = processed_data[1]
            
            comparison_result = compare_dataframes(df1, df2, date1, date2)
            
            if comparison_result:
                diff_df, differences_info = comparison_result
                
                # Criar aba de diferenças
                ws_diff = writer.book.create_sheet("Diferenças")
                
                # Adicionar título
                ws_diff['A1'] = f"Diferenças entre {date1} e {date2}"
                ws_diff.merge_cells(f'A1:{get_column_letter(len(diff_df.columns))}1')
                ws_diff['A1'].font = title_font
                ws_diff['A1'].alignment = title_alignment
                
                # Criar tabela de diferenças
                create_excel_table(ws_diff, diff_df, start_row=3, table_name="Tabela_Diferencas")
                
                # Aplicar formatação nas diferenças
                red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
                
                for diff_info, row_idx in differences_info:
                    excel_row = row_idx + 3  # Ajuste para a posição correta
                    
                    if diff_info is True:  # Linha inteira diferente
                        for col in range(1, len(diff_df.columns) + 1):
                            ws_diff.cell(row=excel_row, column=col).fill = red_fill
                    else:  # Apenas colunas específicas diferentes
                        for col_name in diff_info:
                            col_idx = diff_df.columns.get_loc(col_name) + 1
                            ws_diff.cell(row=excel_row, column=col_idx).fill = red_fill
                
                print(f"\nAba 'Diferenças' criada com {len(diff_df)} linhas diferentes")
        
        # Salvar o arquivo Excel
        writer.close()
        print(f"\nArquivo Excel gerado com sucesso: {output_excel_path}")
    
    except Exception as e:
        print(f"Erro durante o processamento: {str(e)}")

def pdfs_to_excel_with_sheets(pdf_paths, output_excel_path=None):
    """Converte múltiplos PDFs para um único Excel com abas diferentes e tabelas formatadas"""
    try:
        if not pdf_paths:
            print("Nenhum arquivo PDF fornecido.")
            return
        
        if output_excel_path is None:
            # Usa o diretório do primeiro PDF como base
            output_excel_path = os.path.join(os.path.dirname(pdf_paths[0]), "Tabelas_Consolidadas.xlsx")
        
        # Criar um writer Excel
        writer = pd.ExcelWriter(output_excel_path, engine='openpyxl')
        
        # Processar todos os PDFs primeiro
        processed_data = []
        for pdf_path in pdf_paths:
            df, date, filename = process_pdf_to_dataframe(pdf_path)
            if df is not None:
                processed_data.append((df, date, filename, pdf_path))
        
        # Escrever cada PDF em uma aba separada
        for df, date, filename, pdf_path in processed_data:
            sheet_name = date if date else sanitize_sheet_name(filename)
            
            # Se a data já existir como aba, adiciona um sufixo
            original_sheet_name = sheet_name
            counter = 1
            while sheet_name in writer.book.sheetnames:
                sheet_name = f"{original_sheet_name}_{counter}"
                counter += 1
            
            # Escrever no Excel (sem índice)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Acessar a planilha
            ws = writer.book[sheet_name]
            
            # Inserir linha de título
            ws.insert_rows(1)
            ws['A1'] = f"{filename} - {date.replace('_', ' ') if date else filename}"
            ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
            
            # Configurar estilo do título
            title_font = Font(bold=True, size=12)
            title_alignment = Alignment(horizontal='center')
            
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = title_font
                    cell.alignment = title_alignment
            
            # Criar tabela formatada com filtros (começa na linha 3 devido ao título)
            create_excel_table(ws, df, start_row=3, table_name=f"Tabela_{sheet_name}")
            
            print(f"Processado: {pdf_path} -> aba '{sheet_name}' ({len(df)} linhas)")
        
        # Adicionar aba de diferenças se houver exatamente 2 PDFs
        if len(processed_data) == 2:
            df1, date1, filename1, _ = processed_data[0]
            df2, date2, filename2, _ = processed_data[1]
            
            comparison_result = compare_dataframes(df1, df2, date1, date2)
            
            if comparison_result:
                diff_df, differences_info = comparison_result
                
                # Escrever o dataframe de diferenças
                diff_df.to_excel(writer, sheet_name="Diferenças", index=False)
                ws_diff = writer.book["Diferenças"]
                
                # Configurar cabeçalho
                ws_diff.insert_rows(1)
                ws_diff['A1'] = f"Diferenças entre {date1} e {date2}"
                ws_diff.merge_cells(f'A1:{get_column_letter(len(diff_df.columns))}1')
                
                for row in ws_diff.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = title_font
                        cell.alignment = title_alignment
                
                # Aplicar formatação nas diferenças
                red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
                
                for diff_info, row_idx in differences_info:
                    # Ajustar o índice para a linha correta (cabeçalho + 2 linhas adicionais)
                    excel_row = row_idx + 3  # 1 cabeçalho + 1 linha de título + 1 base 1
                    
                    if diff_info is True:  # Linha inteira diferente
                        for col in range(1, len(diff_df.columns) + 1):
                            ws_diff.cell(row=excel_row, column=col).fill = red_fill
                    else:  # Apenas colunas específicas diferentes
                        for col_name in diff_info:
                            col_idx = diff_df.columns.get_loc(col_name) + 1  # +1 para coluna base 1
                            ws_diff.cell(row=excel_row, column=col_idx).fill = red_fill
                
                # Criar tabela formatada com filtros para as diferenças
                create_excel_table(ws_diff, diff_df, start_row=3, table_name="Tabela_Diferencas")
                
                print(f"\nAba 'Diferenças' criada com {len(diff_df)} linhas diferentes")
        
        # Salvar o arquivo Excel
        writer.close()
        print(f"\nArquivo Excel gerado com sucesso: {output_excel_path}")
    
    except Exception as e:
        print(f"Erro durante o processamento: {str(e)}")

# Exemplo de uso - processa todos os PDFs no diretório
pdf_directory = r"C:\Users\win11\Downloads\Tabelas de ofertas"
pdf_files = [os.path.join(pdf_directory, f) for f in os.listdir(pdf_directory) if f.endswith('.pdf')]

if pdf_files:
    pdfs_to_excel_with_sheets(pdf_files)
else:
    print("Nenhum arquivo PDF encontrado no diretório especificado.")